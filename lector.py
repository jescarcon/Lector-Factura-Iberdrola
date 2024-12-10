import os
import re
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

def extract_text_from_pdf(pdf_path):
    """Lee el texto de un archivo PDF utilizando pdfplumber."""
    full_text = ""
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            full_text += page.extract_text() + "\n"
    return full_text

def extract_invoice_data(text):
    """Extrae los datos relevantes de la factura incluyendo el resumen de consumo."""
    direccion_suministro = extract_with_pattern(r"Dirección de suministro\s*:\s*(.+)", text)
    
    # Extraer resumen de facturación y consumo
    resumen_consumo = extract_consumo_data(text)
    
    data = {
        "Periodo de facturación": extract_with_pattern(r"Periodo de facturación (\d{2}/\d{2}/\d{4} - \d{2}/\d{2}/\d{4})", text),
        "Número de factura": extract_with_pattern(r"Número de factura (\d+)", text),
        "Fecha de emisión de factura": extract_with_pattern(r"Fecha de emisión de factura (.+)", text),
        "Fecha prevista de cargo": extract_with_pattern(r"Fecha prevista de cargo (.+)", text),
        "Factura con lectura real": extract_flag(r"Factura con lectura real", text),
        "Titular": extract_with_pattern(r"Titular (.+)", text),
        "CIF titular": extract_with_pattern(r"CIF titular (\w+)", text),
        "Referencia contrato suministro": extract_with_pattern(r"Referencia contrato suministro (\d+)", text),
        "Total importe factura": extract_with_pattern(r"TOTAL IMPORTE FACTURA: ([\d,\.]+ €)", text),
        "Dirección de suministro": direccion_suministro,  # Añadir extracción de la dirección de suministro
        "Resumen de Consumo": resumen_consumo,  # Añadir el resumen de consumo
    }
    return data

def extract_consumo_data(text):
    """Extrae el resumen de consumo (energía, servicios, total)."""
    consumo_data = {
        "Energía": extract_with_pattern(r"ENERGÍA\s*([\d,\.]+ €)", text),
        "Servicios y otros conceptos": extract_with_pattern(r"SERVICIOS Y OTROS CONCEPTOS\s*([\d,\.]+ €)", text),
        "Total a pagar": extract_with_pattern(r"TOTAL A PAGAR\s*([\d,\.]+ €)", text),
    }
    return consumo_data

def extract_with_pattern(pattern, text):
    """Extrae datos usando expresiones regulares."""
    match = re.search(pattern, text)
    return match.group(1) if match else None

def extract_flag(pattern, text):
    """Devuelve True si el patrón se encuentra en el texto, de lo contrario False."""
    return bool(re.search(pattern, text))

def write_to_excel(data_list, output_file):
    """Escribe los datos extraídos en un archivo Excel con cabeceras organizadas."""
    wb = Workbook()
    ws = wb.active

    current_row = 1
    for file_data in data_list:
        filename = file_data["filename"]
        invoice_data = file_data["data"]

        # Añadir la cabecera del fichero
        ws.append([f"Fichero: {filename}"])
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
        ws.cell(row=current_row, column=1).alignment = Alignment(horizontal="left")
        current_row += 1

        # Añadir la subsección de datos de factura
        ws.append(["Datos de Factura"])
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
        ws.cell(row=current_row, column=1).alignment = Alignment(horizontal="left")
        current_row += 1

        # Añadir los datos de la factura
        for key, value in invoice_data.items():
            if key != "Resumen de Consumo" and key != "Dirección de suministro":
                ws.append([key, value])
                current_row += 1

        # Añadir la subsección de resumen de consumo
        ws.append(["Resumen de Consumo"])
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
        ws.cell(row=current_row, column=1).alignment = Alignment(horizontal="left")
        current_row += 1

        # Añadir los datos del resumen de consumo
        resumen_consumo = invoice_data.get("Resumen de Consumo", {})
        for key, value in resumen_consumo.items():
            ws.append([key, value])
            current_row += 1

        # Añadir la subsección de dirección de suministro
        ws.append(["Dirección de suministro"])
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
        ws.cell(row=current_row, column=1).alignment = Alignment(horizontal="left")
        current_row += 1

        # Añadir la dirección de suministro
        direccion_suministro = invoice_data.get("Dirección de suministro", "")
        ws.append([direccion_suministro])
        current_row += 1

        # Añadir una línea en blanco entre facturas
        ws.append([])
        current_row += 1

    wb.save(output_file)
    print(f"Datos guardados en {output_file}")

def write_simple_excel(data_list, output_file):
    """Escribe los datos en un archivo Excel simplificado con columnas planas."""
    wb = Workbook()
    ws = wb.active

    # Crear encabezados
    headers = ["Archivo"]
    example_data = data_list[0]["data"] if data_list else {}
    headers.extend(list(example_data.keys()))
    if "Resumen de Consumo" in headers:
        headers.remove("Resumen de Consumo")
        resumen_keys = list(example_data.get("Resumen de Consumo", {}).keys())
        headers.extend(resumen_keys)
    ws.append(headers)

    # Escribir filas de datos
    for file_data in data_list:
        row = [file_data["filename"]]
        invoice_data = file_data["data"]

        # Añadir datos planos
        for key in headers[1:]:
            if key in invoice_data:
                row.append(invoice_data[key])
            elif key in invoice_data.get("Resumen de Consumo", {}):
                row.append(invoice_data["Resumen de Consumo"].get(key, ""))
            else:
                row.append("")

        ws.append(row)

    wb.save(output_file)
    print(f"Datos guardados en formato simplificado en {output_file}")

def process_pdfs_in_folder(folder_path, output_excel, simple_excel):
    """Procesa todos los archivos PDF en la carpeta especificada y guarda los datos en dos archivos Excel."""
    all_data = []

    # Recorrer todos los archivos en la carpeta
    for filename in os.listdir(folder_path):
        if filename.endswith(".pdf"):
            pdf_path = os.path.join(folder_path, filename)
            print(f"Procesando: {pdf_path}")
            text = extract_text_from_pdf(pdf_path)
            invoice_data = extract_invoice_data(text)
            all_data.append({"filename": filename, "data": invoice_data})

    # Escribir los datos en ambos formatos
    write_to_excel(all_data, output_excel)
    write_simple_excel(all_data, simple_excel)

# Ruta a la carpeta con los archivos PDF y archivos Excel de salida
folder_path = "ficheros"  # Cambia a la ruta de tu carpeta con los PDFs
output_excel = "datos_de_facturas.xlsx"
simple_excel = "datos_simplificados.xlsx"

# Procesar los PDFs en la carpeta y guardar los datos en ambos Excel
process_pdfs_in_folder(folder_path, output_excel, simple_excel)
